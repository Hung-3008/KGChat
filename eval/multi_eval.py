import requests
import time
from datasets import Dataset
import pandas as pd
import os
import sys
import logging
from pydantic import BaseModel, Field
import asyncio
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from dotenv import load_dotenv
import re
from typing import Dict, Any


sys.path.append("./app")
from backend.llm.gemini_client import GeminiClient
from backend.db.neo4j_client import Neo4jClient
from backend.db.vector_db import VectorDBClient
from backend.llm.ollama_client import OllamaClient
from qdrant_client import QdrantClient
from backend.core.retrieval.kg_query_processor import run_query


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("multi_eval.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class EvaluationResult(BaseModel):
    is_correct: int = Field()

class TranslationResult(BaseModel):
    translated_text: str = Field()

class GeminiRateLimiter:
    """Rate limiter for Gemini API calls (10 requests per minute)"""
    def __init__(self, max_requests_per_minute=10):
        self.max_requests = max_requests_per_minute
        self.requests = []
        self.lock = asyncio.Lock()
    
    async def wait_if_needed(self):
        """Wait if we're approaching rate limit"""
        async with self.lock:
            now = time.time()
            # Remove requests older than 1 minute
            self.requests = [req_time for req_time in self.requests if now - req_time < 60]
            
            if len(self.requests) >= self.max_requests:
                # Calculate how long to wait
                oldest_request = min(self.requests)
                wait_time = 61 - (now - oldest_request)  # Wait until oldest request is > 60 seconds old
                if wait_time > 0:
                    logger.info(f"Rate limit approaching, waiting {wait_time:.1f} seconds...")
                    await asyncio.sleep(wait_time)
                    # Clean up again after waiting
                    now = time.time()
                    self.requests = [req_time for req_time in self.requests if now - req_time < 60]
            
            # Record this request
            self.requests.append(now)

class GeminiKeyManager:
    def __init__(
        self,
        current_key_index: int = 1,
        max_keys: int = 6,
        key_pattern: str = "GEMINI_API_KEY_{}"
    ):
        self.current_key_index = current_key_index
        self.max_keys = max_keys
        self.key_pattern = key_pattern
        self.logger = logging.getLogger(__name__)
        self.rate_limiter = GeminiRateLimiter()
    
    def get_current_key(self) -> str:
        key_name = self.key_pattern.format(self.current_key_index)
        api_key = os.getenv(key_name)
        
        if not api_key:
            self.logger.warning(f"API key {key_name} not found in environment variables")
            return None
            
        return api_key
    
    def rotate_key(self) -> str:
        for _ in range(self.max_keys):
            self.current_key_index = (self.current_key_index % self.max_keys) + 1
            
            key_name = self.key_pattern.format(self.current_key_index)
            api_key = os.getenv(key_name)
            
            if api_key:
                self.logger.info(f"Rotated to API key {key_name}")
                # Reset rate limiter for new key
                self.rate_limiter = GeminiRateLimiter()
                return api_key
        
        self.logger.error("No valid API keys found after trying all options")
        return None

class ClientManager:
    """Singleton để quản lý các client connections"""
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super(ClientManager, cls).__new__(cls)
            cls._instance.initialized = False
        return cls._instance
    
    async def initialize(self, gemini_key_manager):
        """Initialize all backend clients"""
        if self.initialized:
            return self.clients
            
        load_dotenv()
        
        # Initialize Neo4j client
        neo4j_uri = os.getenv('NEO4J_URI')
        neo4j_username = os.getenv('NEO4J_USERNAME')
        neo4j_password = os.getenv('NEO4J_PASSWORD')
        
        self.neo4j_client = Neo4jClient(
            uri=neo4j_uri,
            username=neo4j_username,
            password=neo4j_password
        )
        
        # Initialize Vector DB client
        qdrant_host = os.getenv("QDRANT_HOST", "localhost")
        qdrant_port = int(os.getenv("QDRANT_PORT", "6333"))
        
        self.vector_db_client = VectorDBClient(
            host=qdrant_host,
            port=qdrant_port
        )
        
        # Initialize Qdrant client
        self.qdrant_client = QdrantClient(
            host=qdrant_host, 
            port=qdrant_port
        )
        
        # Initialize Ollama client
        ollama_host = os.getenv("OLLAMA_HOST", "http://localhost:11434")
        embedding_model = os.getenv("OLLAMA_EMBEDDING_MODEL", "mxbai-embed-large")
        
        self.ollama_client = OllamaClient(
            host=ollama_host,
            embedding_model=embedding_model
        )
        
        # Initialize Gemini client with key manager
        gemini_api_key = gemini_key_manager.get_current_key()
        self.gemini_client = GeminiClient(
            api_key=gemini_api_key, 
            model_name="gemini-2.0-flash"
        )
        self.gemini_key_manager = gemini_key_manager
        
        # Verify connectivity
        if not await self.neo4j_client.verify_connectivity():
            logger.error("Failed to connect to Neo4j database")
            raise Exception("Neo4j connection failed")
            
        logger.info("All backend clients initialized successfully")
        
        self.clients = {
            "neo4j_client": self.neo4j_client,
            "vector_db_client": self.vector_db_client,
            "ollama_client": self.ollama_client,
            "gemini_client": self.gemini_client,
            "qdrant_client": self.qdrant_client
        }
        
        self.initialized = True
        return self.clients
    
    def rotate_gemini_key(self):
        """Rotate Gemini API key when needed"""
        new_key = self.gemini_key_manager.rotate_key()
        if new_key:
            self.gemini_client = GeminiClient(
                api_key=new_key, 
                model_name="gemini-2.0-flash"
            )
            self.clients["gemini_client"] = self.gemini_client
            return self.gemini_client
        return None
    
    async def close(self):
        """Close all client connections"""
        if hasattr(self, 'neo4j_client'):
            await self.neo4j_client.close()
        self.initialized = False

async def translate_chinese_to_english(gemini_client, key_manager, text, max_retries=3):
    translation_prompt = f"""
    Please translate the following Chinese text to English. 
    The text may contain medical or diabetes-related terminology.
    Provide only the English translation without any additional explanation.
    
    Chinese text:
    {text}
    """
    
    for attempt in range(max_retries):
        try:
            # Wait for rate limiting
            await key_manager.rate_limiter.wait_if_needed()
            
            logger.debug(f"Translating text (attempt {attempt + 1}/{max_retries})")
            
            response = await gemini_client.generate(
                prompt=translation_prompt,
                format=TranslationResult
            )
            
            if isinstance(response, list) and len(response) > 0:
                translated_text = response[0].translated_text
                logger.debug(f"Translation successful: {text[:50]}... -> {translated_text[:50]}...")
                return translated_text
            else:
                logger.warning(f"No translation received on attempt {attempt + 1}")
                
        except Exception as e:
            error_str = str(e).lower()
            logger.error(f"Translation error on attempt {attempt + 1}/{max_retries}: {error_str}")
            
            # Check if it's a rate limit or quota error
            if any(keyword in error_str for keyword in ["resource_exhausted", "rate limit", "429", "quota"]):
                new_key = key_manager.rotate_key()
                if new_key:
                    gemini_client.api_key = new_key
                    os.environ["GEMINI_API_KEY"] = new_key
                    logger.info(f"Switched to API key index {key_manager.current_key_index} for translation")
                else:
                    logger.error("No more API keys available for translation")
                    if attempt == max_retries - 1:
                        break
            
            if attempt < max_retries - 1:
                await asyncio.sleep(3)  # Wait longer for translation retries
    
    logger.error(f"Failed to translate text after {max_retries} attempts: {text[:100]}...")
    return text  # Return original text if translation fails

async def ask_question_direct(question, clients, key_manager, response_type="concise", max_retries=3):
    """
    Directly use backend services instead of API calls
    """
    guided_question = f"""
    This is a multiple-choice question about diabetes. Please read the question carefully and select the most appropriate answer option (A, B, C, or D).

    IMPORTANT INSTRUCTIONS:
    - Provide ONLY the letter and text of your selected answer (e.g., "B) Family history of diabetes")
    - Do NOT include any explanations, reasoning, or additional text
    - Do NOT restate the question
    - Just directly state which option is correct

    QUESTION:
    {question}
    """
    
    conversation_history = []
    
    for attempt in range(max_retries):
        try:
            logger.debug(f"Processing question (attempt {attempt + 1}/{max_retries})")
            
            # Use the backend kg_query_processor directly
            result = await run_query(
                query=guided_question,
                conversation_history=conversation_history,
                clients=clients,
                grounding=False,  # Set to True if you want web grounding
                language="English"
            )
            
            if result and "response" in result:
                return result["response"]
            else:
                logger.warning(f"No response received on attempt {attempt + 1}")
                
        except Exception as e:
            error_str = str(e).lower()
            logger.error(f"Error on attempt {attempt + 1}/{max_retries}: {error_str}")
            
            # Check if it's a rate limit or quota error
            if any(keyword in error_str for keyword in ["resource_exhausted", "rate limit", "429", "quota"]):
                # Try to rotate the Gemini API key
                client_manager = ClientManager()
                new_client = client_manager.rotate_gemini_key()
                if new_client:
                    clients["gemini_client"] = new_client
                    logger.info(f"Rotated to new Gemini API key (index: {key_manager.current_key_index})")
                    await asyncio.sleep(3)  # Wait before retry
                else:
                    logger.error("No more valid Gemini API keys available")
                    if attempt == max_retries - 1:
                        break
            else:
                # For other errors, wait a bit and retry
                await asyncio.sleep(2)
                
        if attempt < max_retries - 1:
            await asyncio.sleep(3)  # Wait between retries
    
    logger.error(f"Failed to get response after {max_retries} attempts")
    return "Error: Failed to get response from backend services"

async def evaluate_answer(gemini_client, key_manager, question, correct_answer, model_answer):
    """
    Evaluate if model's answer matches the correct answer
    """
    prompt = f"""
    You are an evaluation system for multiple-choice answers. 
    
    Question:
    {question}
    
    Correct answer: {correct_answer}
    
    Model's answer: {model_answer}
    
    Your task is to determine if the model's answer correctly identifies the same answer choice as the correct answer.
    
    Rules:
    - The model might not format its answer exactly like the correct answer
    - The model might give additional explanations
    - Focus on whether the model correctly identifies the same option (A, B, C, or D)
    - Return 1 if the model's answer correctly identifies the same option as the correct answer
    - Return 0 if the model's answer does not correctly identify the same option
    
    Do not provide explanations, just evaluate and return the binary result.
    """
    
    max_retries = 3
    
    for attempt in range(max_retries):
        try:
            # Wait for rate limiting
            await key_manager.rate_limiter.wait_if_needed()
            
            response = await gemini_client.generate(
                prompt=prompt,
                format=EvaluationResult
            )
            
            if isinstance(response, list) and len(response) > 0:
                result = response[0].is_correct
            else:
                result = 0                
            return result
            
        except Exception as e:
            logger.error(f"Error during evaluation (attempt {attempt+1}/{max_retries}): {str(e)}")
            
            if attempt < max_retries - 1:
                new_key = key_manager.rotate_key()
                if new_key:
                    gemini_client.api_key = new_key
                    os.environ["GEMINI_API_KEY"] = new_key
                    logger.info(f"Switched to API key index {key_manager.current_key_index} for next attempt")
                else:
                    logger.error("No more API keys available to try")
                    break
                
                await asyncio.sleep(3)  # Longer wait for evaluation retries
    
    logger.warning("All evaluation attempts failed")
    return 0

def create_excel_with_formatting(df, output_path):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Evaluation Results')
    
    workbook = writer.book
    worksheet = writer.sheets['Evaluation Results']
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    incorrect_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Adjust column widths
    worksheet.column_dimensions['A'].width = 15  # question
    worksheet.column_dimensions['B'].width = 50  # original_question (Chinese)
    worksheet.column_dimensions['C'].width = 50  # translated_question (English)
    worksheet.column_dimensions['D'].width = 25  # correct_answer
    worksheet.column_dimensions['E'].width = 25  # model_answer
    worksheet.column_dimensions['F'].width = 15  # is_correct
    
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        
    # Color code the is_correct column
    is_correct_col = df.columns.get_loc('is_correct') + 1
    for row_num, value in enumerate(df['is_correct'], 2):
        cell = worksheet.cell(row=row_num, column=is_correct_col)
        if value == 1:
            cell.fill = correct_fill
            cell.value = "1"
        else:
            cell.fill = incorrect_fill
            cell.value = "0"
    
    writer.close()
    
    logger.info(f"Excel file created at {output_path} with formatting")

async def process_dataset(dataset, clients, gemini_client, key_manager, num_samples=None, batch_size=100, save_prefix='diabetes_mc_results'):
    results = []
    
    total_examples = len(dataset['query'])  # Changed from 'input' to 'query'
    examples_to_process = min(total_examples, num_samples) if num_samples else total_examples
    
    logger.info(f"Starting processing of {examples_to_process} questions")
    
    # Create output directories
    now = datetime.now()
    date_str = now.strftime("%Y_%m_%d")
    time_str = now.strftime("%H_%M_%S")
    
    batch_dir = f"eval/batches/{date_str}/{time_str}"
    os.makedirs(batch_dir, exist_ok=True)
    
    for i in range(examples_to_process):
        original_question = dataset['query'][i]  # Changed from 'input' to 'query'
        correct_answer = dataset['response'][i]  # Changed from 'output' to 'response'
        
        logger.info(f"Processing question {i+1}/{examples_to_process}")
        
        # Translate Chinese question to English
        logger.debug("Translating question from Chinese to English...")
        translated_question = await translate_chinese_to_english(
            gemini_client, 
            key_manager, 
            original_question
        )
        
        # Use translated question for processing
        model_answer = await ask_question_direct(
            question=translated_question, 
            clients=clients, 
            key_manager=key_manager,
            response_type="concise"
        )
        logger.debug(f"Model answered: {model_answer}")
        
        # Evaluate the answer
        is_correct = await evaluate_answer(
            gemini_client, 
            key_manager, 
            translated_question, 
            correct_answer, 
            model_answer
        )
        
        results.append({
            'original_question': original_question,      # Chinese question
            'translated_question': translated_question,  # English translation
            'correct_answer': correct_answer,
            'model_answer': model_answer,
            'is_correct': is_correct
        })
        
        # Save batch periodically
        if (i + 1) % batch_size == 0 or i == examples_to_process - 1:
            batch_number = (i + 1) // batch_size if (i + 1) % batch_size == 0 else ((i + 1) // batch_size) + 1
            batch_df = pd.DataFrame(results)
            batch_filename = f"{batch_dir}/{save_prefix}_batch_{batch_number}.xlsx"
            
            create_excel_with_formatting(batch_df, batch_filename)
            logger.info(f"Saved batch {batch_number} with {len(batch_df)} questions to {batch_filename}")
        
        logger.debug("Waiting to respect rate limits...")
        await asyncio.sleep(20)
    
    logger.info(f"Completed processing all {examples_to_process} questions")
    return pd.DataFrame(results)

async def main():
    logger.info("Starting evaluation process with direct backend integration and translation")
    
    load_dotenv()
    
    key_manager = GeminiKeyManager(current_key_index=4, max_keys=6)
    client_manager = ClientManager()
    
    try:
        clients = await client_manager.initialize(key_manager)
        
        api_key = key_manager.get_current_key()
        if not api_key:
            logger.error("No valid Gemini API key found in environment variables")
            return
        
        gemini_client = GeminiClient(api_key=api_key, model_name="gemini-2.0-flash")
        logger.info(f"Initialized evaluation Gemini client with API key index {key_manager.current_key_index}")
        
        arrow_file_path = "/home/hung/Documents/hung/code/KG_Hung/KGChat/eval/multiple_choice/diabetes_instruct_v8-train.arrow"
        logger.info(f"Loading dataset from {arrow_file_path}")
        
        try:
            dataset = Dataset.from_file(arrow_file_path)
            dataset_dict = dataset.to_dict()
            
            expected_keys = ['query', 'response']
            actual_keys = list(dataset_dict.keys())
            logger.info(f"Dataset keys found: {actual_keys}")
            
            if not all(key in actual_keys for key in expected_keys):
                logger.error(f"Expected keys {expected_keys} but found {actual_keys}")
                return
            
            logger.info(f"Dataset loaded successfully with {len(dataset_dict['query'])} questions")
        except Exception as e:
            logger.error(f"Failed to load dataset: {str(e)}")
            return
        
        now = datetime.now()
        date_str = now.strftime("%Y_%m_%d")
        time_str = now.strftime("%H_%M_%S")
        
        output_dir = f"eval/multiple_choice/{date_str}/{time_str}"
        os.makedirs(output_dir, exist_ok=True)
        
        num_samples = 10 
        if num_samples:
            logger.info(f"Will process {num_samples} samples")
        else:
            logger.info("Processing all samples in the dataset")
        
        results_df = await process_dataset(
            dataset=dataset_dict, 
            clients=clients,
            gemini_client=gemini_client,
            key_manager=key_manager,
            num_samples=num_samples,
            batch_size=10,  
            save_prefix='diabetes_mc_results'
        )
        
        # Create final Excel file with formatting
        final_output_path = f"{output_dir}/diabetes_mc_results_complete.xlsx"
        create_excel_with_formatting(results_df, final_output_path)
        logger.info(f"Saved complete results to {final_output_path}")
        
        # Calculate and log summary statistics
        total_questions = len(results_df)
        correct_count = results_df['is_correct'].sum()
        accuracy = (correct_count / total_questions) * 100 if total_questions > 0 else 0
        
        logger.info("\nResults Summary:")
        logger.info(f"Total questions processed: {total_questions}")
        logger.info(f"Correct answers: {correct_count} ({accuracy:.2f}%)")
        
        # Create summary sheet
        summary_df = pd.DataFrame([{
            'Total Questions': total_questions,
            'Correct Answers': correct_count,
            'Accuracy': f"{accuracy:.2f}%",
            'Date': pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
            'Translation Enabled': True
        }])
        
        summary_file = f"{output_dir}/diabetes_mc_results_summary.xlsx"
        summary_df.to_excel(summary_file, index=False)
        logger.info(f"Saved summary statistics to {summary_file}")
        
    except Exception as e:
        logger.error(f"Error in main process: {str(e)}", exc_info=True)
    finally:
        # Clean up connections
        await client_manager.close()
        logger.info("Closed all backend connections")

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except Exception as e:
        logger.error(f"Unhandled exception: {str(e)}", exc_info=True)